import tkinter # import de tkinter
from tkinter import *
from tkinter import ttk
from tkinter import messagebox
import pandas as pd
import recherche_produit
import update
import openpyxl
import calendar
from babel.dates import format_date, parse_date, get_day_names, get_month_names
from babel.numbers import *
from tkcalendar import DateEntry
from tkcalendar import Calendar
import datetime
from datetime import date
import time
import requests
import os
import zipfile
import webbrowser
import numpy

#"""
#Pour compiler le programme pour Windows : décommenter les 2 lignes ci-dessous :
#import pyi_splash
#pyi_splash.close()
#
#"""
#    - Décommenter les lignes ci-dessus (import pyi_splash ; pyi_splash.close() )
#    - Décommenter la ligne '#fenetre.iconbitmap('icone.ico')' (pour avoir l'icone)
#    - Lancer "cmd" en mode Administrateur
#    - Se placer dans le dossier qui contient 'main.py' (en tapant par exemple : cd C:\Users\luc\Desktop\v25 )
#    - Lancer la commande suivante : pyinstaller main.py --onefile -w --splash splashscreen.png
#    - Le fichier .exe se trouve dans le dossier 'dist'.
#    - Le copier dans le dossier principal pour pouvoir lancer le programme
#    - On peut supprimer 'main.py', 'recherche_produit.py' et 'update.py'
#"""

fenetre = tkinter.Tk ()                         # création de la fenêtre principale

# test theme
style = ttk.Style(fenetre)
fenetre.tk.call("source", "azure-theme/azure.tcl")
fenetre.tk.call("set_theme", "light")
# https://github.com/rdbende/Azure-ttk-theme

fenetre.geometry("750x600")                     # taille de fenêtre
fenetre.title("IFT - v0.2 - Jean-Pierre & Luc Drouillet (2022)")                 # titre
fenetre.resizable(width=True, height=True)    # fenetre redimensionnable ?
# fenetre.iconbitmap('icone.ico')

# Année en cours
annee_en_cours = int(datetime.datetime.now().date().strftime("%Y"))

# Onglets
onglets = ttk.Notebook(fenetre,width=750,height=600)   # Création du système d'onglets
onglets.grid(row=14,column=1)
onglet1 = ttk.Frame(onglets)       # Ajout de l'onglet 1
onglet1.grid(row=15,column=0)
onglet2 = ttk.Frame(onglets)       # Ajout de l'onglet 2
onglet2.grid(row=15,column=1)
onglet4 = ttk.Frame(onglets)       # Ajout de l'onglet 4
onglet4.grid(row=15,column=3)
onglet5 = ttk.Frame(onglets)       # Ajout de l'onglet 3
onglet5.grid(row=15,column=4)
onglet3 = ttk.Frame(onglets)
onglet3.grid(row=15, column=2)
onglets.add(onglet1, text='Traitement')      # Nom de l'onglet 1
onglets.add(onglet2, text='Parcelles')# Nom de l'onglet 2
onglets.add(onglet3, text='Traitements réalisés')
onglets.add(onglet4, text='Récapitulatif')   # Nom de l'onglet 3
#onglets.add(onglet5, text='Update')   # Nom de l'onglet 3


# Création des variables : Certaines sont devenues inutiles TODO à nettoyer
dictionnaire_parcelles = {}     # Dictionnaire où sont stockées les noms de parcelles et leur superficie
dictionnaire_confusion_sexuelle = {}     # Dictionnaire où sont stockées les noms de parcelles et leur superficie en confusion sexuelle
liste=[]                        # Liste des produits qui correspondent à la recherche de l'utilisateur
herbicide=0
produit_choisi = StringVar()    # Produit choisi dans la liste
liste_doses_pour_le_produit=[]  # Liste des doses possibles pour le produit choisi
dose_autorisee=StringVar()      # Dose autorisée pour le produit en cours
unite=StringVar()               # Unité pour la dose autorisée
application=StringVar()         # Nombre de passages maximum pour le produit en cours
dose_choisie = StringVar()      # Dose choisie dans la liste
IFT=StringVar()                 # Valeur de l'IFT
parcelle_choisie=StringVar()    # Parcelle choisie (uniquement si dictionnaire de parcelles)
parcelle_a_supprimer=StringVar()# Parcelle de la liste "à supprimer" onglet2
IFT_total_parcelle=StringVar()  # Quand l'IFT total de la parcelle a été calculé
label_passages_restants=StringVar()  # Si limite de passages, afficher ce qu'il reste pour la parcelle/produit
IFT_base_de_donnees=StringVar() # Pour calculer l'IFT sur toutes les parcelles (IFT par parcelle / nb de parcelles)
resultat_mise_a_jour=StringVar()# Pour dire si la mise à jour a été faite
utilisation_choisie=StringVar() # Pour dire quelle est l'utilisation du produit par l'utilisateur
superficie_confusion_sexuelle=StringVar() # Si la parcelle en confusion sexuelle indiquer la superficie
produit_herbicide=StringVar() # Si produit herbicide, afficher une liste pour choisir "pré/post-levée"
# Variables onglet récap
liste_parcelles_recap=StringVar()
IFT_total_recap=StringVar()
IFT_classique_recap=StringVar()
IFT_biocontrole_recap=StringVar()
IFT_mildiou_recap=StringVar()
IFT_oidium_recap=StringVar()
IFT_botrytis_recap=StringVar()
IFT_autres_fongicides_recap=StringVar()
IFT_fongicides_total_recap=StringVar()
IFT_confusion_sexuelle_recap=StringVar()
IFT_acariens_recap=StringVar()
IFT_insecticide_recap=StringVar()
IFT_hors_herbicides_recap=StringVar()
IFT_herbicides_prelevee_recap=StringVar()
IFT_herbicides_postlevee_recap=StringVar()
IFT_herbicides_recap=StringVar()
IFT_autres_acaricides_recap=StringVar()
# Gestion résistances
gestion_phenylpyrroles=StringVar()
gestion_ANP=StringVar()
gestion_IBS3=StringVar()
gestion_SDHI=StringVar()
gestion_CAA=StringVar()
gestion_zoxamide=StringVar()
gestion_Qil=StringVar()
gestion_qosi=StringVar()
gestion_fluopicolide=StringVar()
gestion_oxathiapiproline=StringVar()
gestion_anilides=StringVar()
gestion_cymoxanil=StringVar()
gestion_qoicontact=StringVar()
gestion_spiroxamine=StringVar()
gestion_APK=StringVar()
gestion_SDHI_fluopyram=StringVar()
gestion_SDHI_boscalid=StringVar()
gestion_SDHI_fluxapyroxad=StringVar()
gestion_IBS1=StringVar()
gestion_AZN=StringVar()
gestion_cyflufenamid=StringVar()
periode_indiquee = 0
# Cases à cocher
resume_IFT= IntVar(value=1)
resume_gestion_resistances= IntVar(value=1)
resume_traitements= IntVar(value=1)

# Récupérer les parcelles et leur superficie stockées dans le fichier 'liste_parcelles.xlsx'
fichier_parcelles = pd.read_excel('liste_parcelles.xlsx')
if len(fichier_parcelles) > 0 :
    for ligne in range(len(fichier_parcelles)):
        recuperer_nom_parcelle=fichier_parcelles['nom_parcelle'].iloc[ligne]
        recuperer_taille_parcelle=fichier_parcelles['taille_parcelle'].iloc[ligne]
        dictionnaire_parcelles[recuperer_nom_parcelle]=recuperer_taille_parcelle
else :
    dictionnaire_parcelles={}
    dictionnaire_confusion_sexuelle = {}

# Déclaration des fonctions :

def a_propos():
    messagebox.showinfo("À propos - Luc & Jean-Pierre Drouillet","NOM_DU_PROGRAMME est un programme de suivi de traitements, adapté à la viticulture. \n\nCe logiciel est écrit par Luc Drouillet et Jean-Pierre Drouillet. \n\nSi vous rencontrez des problèmes avec IFT, vous pouvez nous contacter par le biais de ce site internet : https://XXXXXXXXXX")

def tkquit():
    fenetre.destroy()

# Regrouper les 4 fonctions suivantes en une seule TODO
def changer_virgule_en_point_surface_traitee():
    text = surface_traitee.get()
    text = text.replace(',', '.')
    surface_traitee.delete(0, 'end')
    surface_traitee.insert(0, text)

def changer_virgule_en_point_quantite_traitee():
    text = quantite_traitee.get()
    text = text.replace(',', '.')
    quantite_traitee.delete(0, 'end')
    quantite_traitee.insert(0, text)

def changer_virgule_en_point_entrer_taille_de_parcelle():
    text = entrer_taille_de_parcelle.get()
    text = text.replace(',', '.')
    entrer_taille_de_parcelle.delete(0, 'end')
    entrer_taille_de_parcelle.insert(0, text)

def changer_virgule_en_point_entrer_superficie_confusion():
    text = entrer_superficie_confusion.get()
    text = text.replace(',', '.')
    entrer_superficie_confusion.delete(0, 'end')
    entrer_superficie_confusion.insert(0, text)


def rechercher_produit():                               # Lister les produits qui correspondent à la recherche de l'utilisateur
    produit=nom_produit.get()                           # Stocker ce qu'a recherché l'utilisateur dans la variable "produit"
    liste=recherche_produit.produits_possibles(produit) # Faire appel à la fonction 'produits_possibles()' du script 'recherche_produit'. Stocker les résultats dans la variable 'liste'
    if liste == ["Le produit n'est pas trouvé."] :      # Si la liste des produits trouvés est vide : tout effacer dans le programme
        combobox_liste_produits['values'] = liste
        combobox_liste_produits.current(0)
        combobox_utilisations['values'] = [""]
        combobox_utilisations.current(0)
        dose_autorisee.set("")
        unite.set("")
        application.set("")
        quantite_traitee.delete(0,END)
        IFT.set("")
        label_passages_restants.set("")
        combobox_herbicide.grid_forget()
        produit_herbicide.set("non herbicide")
    elif liste == ["Le produit a été retiré."] :        # Si le produit a été retiré : tout effacer également
        combobox_liste_produits['values'] = liste
        combobox_liste_produits.current(0)
        combobox_utilisations['values'] = [""]
        combobox_utilisations.current(0)
        dose_autorisee.set("")
        unite.set("")
        application.set("")
        quantite_traitee.delete(0,END)
        IFT.set("")
        label_passages_restants.set("")
        combobox_herbicide.grid_forget()
        produit_herbicide.set("non herbicide")
    else :                                              # Sinon c'est que le produit est trouvé et autorisé
        combobox_liste_produits['values'] = liste       # Insérer les produits trouvés dans la liste de produits
        combobox_liste_produits.current(0)              # Afficher directement le premier élément de la liste (0)
        rechercher_meme_dose()                          # Lancer les fonctions recherche des doses et le calcul de l'IFT pour le produit (permet de ne pas avoir à cliquer sur les boutons)

def rechercher_meme_dose():                                                             # Lister les doses possibles pour le produit trouvé
    produit_combobox=produit_choisi.get()                                               # Récupérer le nom du produit dans la liste et le stocker dans la variable "produit_combobox"
    liste_doses_possibles,liste_utilisations=recherche_produit.verifier_si_meme_doses(produit_combobox)    # Stocker les doses possibles pour ce produit dans 'liste_doses_possibles'
    liste_doses_possibles = [x for x in liste_doses_possibles if pd.isnull(x) == False] # Retirer les valeurs "NaN" de la liste s'il y en a
    liste_utilisations_neuve=[]
    for i in range(len(liste_utilisations)):
        liste_utilisations_neuve.append(str(i+1)+str("-")+str(liste_utilisations[i]))
    combobox_utilisations['values'] = liste_utilisations_neuve                                 # Insérer les doses possibles dans la liste des doses
    combobox_utilisations.current(0)                                                    # Afficher directement le premier élément de la liste des doses (0)
    dose_choisie.set(liste_doses_possibles[0])
    rechercher_doses_unite()                                                            # Lancer la fonction qui affiche "dose autorisée", "unité" et "passages max" pour ce produit avec cette dose

def rechercher_doses_unite():                                                                                       # Récupérer la dose autorisée, l'unité, et le nombre de passages max.
    utilisation_en_cours=utilisation_choisie.get()
    index_utilisation_en_cours=str(utilisation_en_cours)[0]
    utilisation_en_cours=str(utilisation_en_cours)[2:]
    produit_combobox=produit_choisi.get()
    liste_doses_possibles,liste_utilisations=recherche_produit.verifier_si_meme_doses(produit_combobox)    # Stocker les doses possibles pour ce produit dans 'liste_doses_possibles'
    liste_doses_possibles = [x for x in liste_doses_possibles if pd.isnull(x) == False] # Retirer les valeurs "NaN" de la liste s'il y en a
    dose_choisie.set(liste_doses_possibles[int(index_utilisation_en_cours)-1])
    dose=dose_choisie.get()
    recherche_produit.choisir_ligne_en_fonction_dose(produit_combobox,dose,utilisation_en_cours)

    dose_autorisee1,unite1,application1,herbicide=recherche_produit.choisir_ligne_en_fonction_dose(produit_combobox,dose,utilisation_en_cours)
    dose_autorisee.set(dose_autorisee1)
    unite.set(unite1)

    # Si produit est HERBICIDE afficher une liste qui contient "pré-levée" / "post-levée"
    if herbicide == 1 :
        combobox_herbicide.grid(row=0,column=1)
        combobox_herbicide.current(0)
        herbicide=0
    else :
        combobox_herbicide.grid_forget()
        produit_herbicide.set("non herbicide")


    if application1 > 0:
        application.set(int(application1))
    else :
        application1=0
        application.set("Pas de limite")
    quantite_traitee.delete(0,END)
    quantite_traitee.insert(0,dose_autorisee.get())
    # ICIICI
    s2.configure(to=float(dose_autorisee.get()))
    s2.set(dose_autorisee.get())
    #
    surface_traitee.delete(0,END)
    surface_traitee.insert(0,dictionnaire_parcelles[parcelle_choisie.get()])
    calcul_IFT()

def calcul_IFT():
    nombre_passages_restants()
    dose_app=quantite_traitee.get()
    surface_tra=surface_traitee.get()
    dose_auto=dose_autorisee.get()
    superficie_parcelle_en_cours=dictionnaire_parcelles[parcelle_choisie.get()]
    try:
        calcul_IFT = round((float(dose_app) * float(surface_tra)) / (float(dose_auto) * float(superficie_parcelle_en_cours)),2) # arrondi à 2 décimales
    except ZeroDivisionError:
        calcul_IFT=""
    IFT.set(calcul_IFT)

def ajouter_parcelle():
    # Ajoute au dictionnaire en cours
    nom_nouvelle_parcelle=entrer_un_nom_de_parcelle.get()
    taille_nouvelle_parcelle=float(entrer_taille_de_parcelle.get())
    if superficie_confusion_sexuelle.get() == "":
        taille_confusion=0
    else :
        taille_confusion=float(superficie_confusion_sexuelle.get())
    if taille_confusion > taille_nouvelle_parcelle:
        messagebox.showinfo("Information","La surface en confusion sexuelle est supérieure à la superficie de la parcelle.")
        return
    entrer_un_nom_de_parcelle.delete(0, END)
    entrer_taille_de_parcelle.delete(0, END)
    entrer_superficie_confusion.delete(0,END)
    dictionnaire_parcelles[nom_nouvelle_parcelle]=taille_nouvelle_parcelle
    dictionnaire_confusion_sexuelle[nom_nouvelle_parcelle]=taille_confusion
    combobox_parcelles['values'] = list(dictionnaire_parcelles.keys())
    liste_parcelles_pour_recap['values']=["Toute l'exploitation"]+list(dictionnaire_parcelles.keys())
    liste_parcelles_pour_recap.current(0)
    liste_parcelles_de_la_base['values'] = list(dictionnaire_parcelles.keys())
    combobox_parcelles.current(0)
    liste_parcelles_de_la_base.current(0)
    parcelle_selectionnee.config(text=str(dictionnaire_parcelles[parcelle_choisie.get()])+" hectare(s)")
    surface_traitee.delete(0,END)
    surface_traitee.insert(0,dictionnaire_parcelles[parcelle_choisie.get()])
    # Manipuler le fichier excel
    fichier_parcelles_openpyxl = openpyxl.reader.excel.load_workbook('liste_parcelles.xlsx')
    sheet_en_cours = fichier_parcelles_openpyxl['liste_parcelles']
    sheet_en_cours.append([nom_nouvelle_parcelle, taille_nouvelle_parcelle, taille_confusion])
    fichier_parcelles_openpyxl.save('liste_parcelles.xlsx')
    fichier_parcelles = pd.read_excel('liste_parcelles.xlsx')
    messagebox.showinfo("Information","La parcelle a bien été ajoutée.")
    infos_IFT_toute_la_base_de_donnees()
    recapitulatif_total()

def supprimer_parcelle():
    # Afficher un message d'avertissement
    avertissement=tkinter.messagebox.askyesno(title="Supprimer la parcelle", message="Êtes-vous sûr de vouloir supprimer cette parcelle de toute la base de données ?")
    if avertissement == True :
        parcelle_suppression=parcelle_a_supprimer.get()
        dictionnaire_parcelles.pop(parcelle_suppression)
        if dictionnaire_confusion_sexuelle :
            dictionnaire_confusion_sexuelle.pop(parcelle_suppression)
        combobox_parcelles['values'] = list(dictionnaire_parcelles.keys())
        liste_parcelles_de_la_base['values'] = list(dictionnaire_parcelles.keys())
        if dictionnaire_parcelles:
            combobox_parcelles.current(0)
            liste_parcelles_de_la_base.current(0)
            parcelle_selectionnee.config(text=str(dictionnaire_parcelles[parcelle_choisie.get()])+" hectare(s)")
            surface_traitee.delete(0,END)
            surface_traitee.insert(0,dictionnaire_parcelles[parcelle_choisie.get()])
            liste_parcelles_pour_recap['values']=["Toute l'exploitation"]+list(dictionnaire_parcelles.keys())
            liste_parcelles_pour_recap.current(0)
        else:
            surface_traitee.delete(0,END)
            quantite_traitee.delete(0,END)
            IFT.set("")
            label_passages_restants.set("")
            combobox_parcelles['values'] = [""]
            liste_parcelles_de_la_base['values'] = [""]
            combobox_parcelles.current(0)
            liste_parcelles_de_la_base.current(0)
            parcelle_selectionnee.config(text=str(""))
            liste_parcelles_pour_recap['values']=["Toute l'exploitation"]+list(dictionnaire_parcelles.keys())
            liste_parcelles_pour_recap.current(0)
            IFT_total_recap.set("")
            IFT_classique_recap.set("")
            IFT_biocontrole_recap.set("")
            IFT_mildiou_recap.set("")
            IFT_oidium_recap.set("")
            IFT_botrytis_recap.set("")
            IFT_autres_fongicides_recap.set("")
            IFT_fongicides_total_recap.set("")
            IFT_confusion_sexuelle_recap.set("")
            IFT_acariens_recap.set("")
            IFT_autres_acaricides_recap.set("")
            IFT_insecticide_recap.set("")
            IFT_hors_herbicides_recap.set("")
            IFT_herbicides_prelevee_recap.set("")
            IFT_herbicides_postlevee_recap.set("")
            IFT_herbicides_recap.set("")
        # Manipuler le fichier excel "liste parcelles"
        fichier_parcelles = pd.read_excel('liste_parcelles.xlsx')
        index_a_supprimer = fichier_parcelles[ fichier_parcelles['nom_parcelle'] == parcelle_suppression ].index
        fichier_parcelles = fichier_parcelles.drop(index_a_supprimer)
        fichier_parcelles.to_excel('liste_parcelles.xlsx', sheet_name='liste_parcelles', index=False)
        fichier_parcelles = pd.read_excel('liste_parcelles.xlsx')
        # Manipuler le fichier excel "IFT"
        fichier_IFT = pd.read_excel('IFT.xlsx')
        index_a_supprimer = fichier_IFT[ fichier_IFT['nom_parcelle'] == parcelle_suppression ].index
        fichier_IFT = fichier_IFT.drop(index_a_supprimer)
        fichier_IFT.to_excel('IFT.xlsx', sheet_name='IFT', index=False)
        fichier_IFT = pd.read_excel('IFT.xlsx')
        infos_IFT_par_parcelle()
        infos_IFT_toute_la_base_de_donnees()
        recapitulatif_total()
        lire_traitements()
    else :
        return

def sauvegarde():
    if produit_choisi.get() == "" :
        messagebox.showinfo("Information","Veuillez sélectionner un produit.")
    elif parcelle_choisie.get() == "":
        messagebox.showinfo("Information","Veuillez indiquer la parcelle traitée.")
    elif surface_traitee.get() == "":
        messagebox.showinfo("Information","Veuillez indiquer la surface traitée.")
    elif quantite_traitee.get() == "" :
        messagebox.showinfo("Information","Veuillez indiquer la dose appliquée pour ce traitement.")
    elif IFT.get() == "" :
        messagebox.showinfo("Information","L'IFT n'a pas pu être calculé.")
    elif float(surface_traitee.get()) > float(dictionnaire_parcelles[parcelle_choisie.get()]):
        messagebox.showinfo("Information","La surface traitée est supérieure à la superficie de la parcelle.")
    elif float(quantite_traitee.get()) > float(dose_autorisee.get()):
        messagebox.showinfo("Information","La quantité de produit traitée est supérieure à la dose réglementaire.")
    else :
        passages_restants=nombre_passages_restants()
        if passages_restants == 0:
            avertissement=messagebox.askyesno("Information","Vous avez atteint le nombre de passages maximal autorisé pour ce produit dans cette parcelle.\nContinuer ?")
            if avertissement == False:
                return
        date_du_jour,nom_parcelle,nom_produit,dose_mise,dose_reglementaire,superficie_traitee,superficie_totale_parcelle,IFT_recupere,passage_numero,utilisation,biocontrole,numero_amm,substances_actives,fonction_produit,unite=recuperer_toutes_les_infos_pour_ce_traitement()
        date_traitement=indiquer_date.get_date()
        date_traitement=date_traitement.strftime('%d/%m/%Y')
        # Ajouter à l'excel 'IFT.xlsx'
        fichier_IFT_openpyxl = openpyxl.reader.excel.load_workbook('IFT.xlsx')
        sheet_en_cours = fichier_IFT_openpyxl['IFT']
        fichier_IFT = pd.read_excel('IFT.xlsx')
        si_herbicide=produit_herbicide.get()
        sheet_en_cours.append([date_du_jour, date_traitement,nom_produit, numero_amm, substances_actives, utilisation ,fonction_produit,si_herbicide,nom_parcelle,dose_mise,dose_reglementaire,unite,superficie_traitee,superficie_totale_parcelle,IFT_recupere,biocontrole,passage_numero])
        fichier_IFT_openpyxl.save('IFT.xlsx')
        fichier_IFT = pd.read_excel('IFT.xlsx')
        messagebox.showinfo("Information","Le traitement a été enregistré.")
        infos_IFT_par_parcelle()
        infos_IFT_toute_la_base_de_donnees()
        nombre_passages_restants()
    recapitulatif_total()
    lire_traitements()

def recuperer_toutes_les_infos_pour_ce_traitement():
        date_du_jour=date.today()
        date_du_jour=date_du_jour.strftime("%d/%m/%Y")      # Date
        nom_parcelle=parcelle_choisie.get()                 # Nom de la parcelle
        nom_produit=produit_choisi.get()                    # Nom du produit
        dose_mise=quantite_traitee.get()                         # Dose appliquée
        dose_reglementaire=dose_autorisee.get()
        unite_dose_reglementaire=unite.get()
        superficie_traitee=surface_traitee.get()            # Surface traitée
        superficie_totale_parcelle=dictionnaire_parcelles[parcelle_choisie.get()]
        IFT_recupere=IFT.get()                              # IFT pour ce traitement
        passage_numero=1                                    # Ajouter un passage pour ce traitement (rentrera dans le calcul des passages autorisés restants)
        utilisation=utilisation_choisie.get()
        utilisation=str(utilisation)[2:]                    # Utilisation choisie
        biocontrole,numero_amm,substances_actives,fonction_produit=recherche_produit.recuperer_infos(nom_produit,utilisation,dose_reglementaire)
        return(date_du_jour,nom_parcelle,nom_produit,dose_mise,dose_reglementaire,superficie_traitee,superficie_totale_parcelle,IFT_recupere,passage_numero,utilisation,biocontrole,numero_amm,substances_actives,fonction_produit,unite_dose_reglementaire)



def infos_IFT_par_parcelle():
    # IFT total pour la parcelle
    IFT_total=0
    IFT_en_cours=0
    parcelle_a_etudier=parcelle_a_supprimer.get()
    fichier_IFT = pd.read_excel('IFT.xlsx') # TODO
    lignes_contenant_parcelle=fichier_IFT[fichier_IFT['nom_parcelle'].str.contains(parcelle_a_etudier, na=False)]
    for ligne in range(len(lignes_contenant_parcelle)):
        IFT_en_cours=lignes_contenant_parcelle['IFT'].iloc[ligne]
        IFT_total=round(IFT_total+IFT_en_cours, 2)
    IFT_total_parcelle.set(IFT_total)

def infos_IFT_toute_la_base_de_donnees():
    # IFT total sur l'ensemble des parcelles
    # formule : (somme (IFT parcelle * taille parcelle)) / (taille exploitation)  <- regarder IFT biocontrole vs classqiue
    IFT_BDD=0
    fichier_IFT = pd.read_excel('IFT.xlsx')
    if len(fichier_IFT) == 0 :
        IFT_base_de_donnees.set(0)
    else :
        taille_totale_exploitation=0
        IFT_parcelle=0
        IFT_en_cours=0
        for parcelle in list(dictionnaire_parcelles.keys()):
            taille_totale_exploitation=float(taille_totale_exploitation)+float(dictionnaire_parcelles[parcelle])
        for parcelle in list(dictionnaire_parcelles.keys()) :
            liste_des_parcelles_dans_IFT=fichier_IFT[fichier_IFT['nom_parcelle'].str.contains(parcelle, na=False)]
            for ligne in range(len(liste_des_parcelles_dans_IFT)):
                IFT_en_cours=liste_des_parcelles_dans_IFT['IFT'].iloc[ligne]
                IFT_parcelle=IFT_parcelle+IFT_en_cours
            IFT_parcelle=float(IFT_parcelle)*float(dictionnaire_parcelles[parcelle])
            IFT_BDD=IFT_BDD+IFT_parcelle
            IFT_en_cours=0
            IFT_parcelle=0
        IFT_BDD=round(IFT_BDD/taille_totale_exploitation,2)
        IFT_base_de_donnees.set(IFT_BDD)

def nombre_passages_restants():
    nombre_passages_autorises=application.get()
    nombre_passages_total=0
    nombre_passages_en_cours=0
    nom_produit_choisi=produit_choisi.get()
    il_reste_des_passages=1
    if nom_produit.get() == "":
        infos_passages=""
        label_passages_restants.set(infos_passages)
        return(il_reste_des_passages)
    if nombre_passages_autorises == "Pas de limite":
        infos_passages=""
        label_passages_restants.set(infos_passages)
        return(il_reste_des_passages)
    else :
        parcelle_a_etudier=parcelle_choisie.get()
        fichier_IFT = pd.read_excel('IFT.xlsx')
        lignes_contenant_parcelle=fichier_IFT[fichier_IFT['nom_parcelle'].str.contains(parcelle_a_etudier, na=False)]
        lignes_contenant_parcelle=lignes_contenant_parcelle[lignes_contenant_parcelle['nom_produit'].str.contains(nom_produit_choisi, na=False)]
        for ligne in range(len(lignes_contenant_parcelle)):
            nombre_passages_en_cours=lignes_contenant_parcelle['passage'].iloc[ligne]
            nombre_passages_total=nombre_passages_total+nombre_passages_en_cours
        passages_restants=int(nombre_passages_autorises)-int(nombre_passages_total)
        if passages_restants >= 0:
            #infos_passages=str("Il reste ") + str(passages_restants) + str(" passages autorisés avec ce produit.")
            infos_passages=str("Il vous reste ") + str(passages_restants) + str(" passages autorisés dans cette parcelle.")
            label_passages_restants.set(infos_passages)  # Si limite de passages, afficher ce qu'il reste pour la parcelle/produit
            if passages_restants == 0:
                il_reste_des_passages=0
                return(il_reste_des_passages)
        else :
            infos_passages=str("Vous avez dépassé le nombre de passages autorisés (") + str(nombre_passages_total) + str(" passages)")
            label_passages_restants.set(infos_passages)  # Si limite de passages, afficher ce qu'il reste pour la parcelle/produit
            il_reste_des_passages=0
            return(il_reste_des_passages)

def mettre_a_jour_bdd():
    chemin_nouvelle_base=str(os.getcwd())+str("/bdd_phyto.csv")
    infos=update.mettre_a_jour()
    if infos == "Votre base de données de produits est déjà à jour." :
        messagebox.showinfo("Information","Votre base de données de produits est déjà la plus récente.")
    else :
        messagebox.showinfo("Information","La base de données de produits a bien été mise à jour !")
        df = pd.read_csv (chemin_nouvelle_base,sep=';')
        resultat_mise_a_jour.set(infos)
        rechercher_produit()

def update_superficie_a_choisir() :
    surface_traitee.delete(0,END)
    surface_traitee.insert(0,dictionnaire_parcelles[parcelle_choisie.get()])
    calcul_IFT()


def recapitulatif_total():
    nom_parcelle = liste_parcelles_recap.get()
    surface_totale_exploitation = 0
    for parcelle in list(dictionnaire_parcelles.keys()) :
        surface_totale_exploitation=float(surface_totale_exploitation)+float(dictionnaire_parcelles[parcelle])

    date_debut = indiquer_date_debut.get()
    date_fin = indiquer_date_fin.get()
    if date_debut >= date_fin:
        periode_indiquee = 0
    else :
        periode_indiquee = 1

    IFT_classique , IFT_biocontrole , IFT_total , IFT_mildiou , IFT_oidium , IFT_botrytis , IFT_autres_fongicides , IFT_fongicide_total , IFT_confusion_sexuelle , IFT_acaricides , IFT_autres_acaricides , IFT_insecticide_total , IFT_hors_herbicides , IFT_herbicide_prelevee, IFT_herbicide_postlevee, IFT_herbicide , cyflufenamid, AZN, IBS1, SDHI_fluxapyroxad, SDHI_boscalid, SDHI_fluopyram, APK, spiroxamine, qoicontact, cymoxanil, anilides, oxathiapiproline, fluopicolide, qosi, Qil, zoxamide, CAA, SDHI, IBS3, ANP, phenylpyrroles, premiere_date_traitement = recherche_produit.recapitulatif(nom_parcelle,surface_totale_exploitation,periode_indiquee,date_debut,date_fin)

    # Set les IFT
    IFT_total_recap.set(IFT_total)
    IFT_classique_recap.set(IFT_classique)
    IFT_biocontrole_recap.set(IFT_biocontrole)
    IFT_mildiou_recap.set(IFT_mildiou)
    IFT_oidium_recap.set(IFT_oidium)
    IFT_botrytis_recap.set(IFT_botrytis)
    IFT_autres_fongicides_recap.set(IFT_autres_fongicides)
    IFT_fongicides_total_recap.set(IFT_fongicide_total)
    IFT_confusion_sexuelle_recap.set(IFT_confusion_sexuelle)
    IFT_acariens_recap.set(IFT_acaricides)
    IFT_autres_acaricides_recap.set(IFT_autres_acaricides)
    IFT_insecticide_recap.set(IFT_insecticide_total)
    IFT_hors_herbicides_recap.set(IFT_hors_herbicides)
    IFT_herbicides_prelevee_recap.set(IFT_herbicide_prelevee)
    IFT_herbicides_postlevee_recap.set(IFT_herbicide_postlevee)
    IFT_herbicides_recap.set(IFT_herbicide)
    # Set la gestion des resistances
    gestion_phenylpyrroles.set(str(phenylpyrroles)+" (1 an/parcelle)")
    gestion_ANP.set(str(ANP)+" (1 an/parcelle) 1 année sur 2")
    gestion_IBS3.set(str(IBS3)+" (1 an/parcelle) 1 année sur 2")
    gestion_SDHI.set(str(SDHI)+" (1 an/parcelle) sauf boscalid 1 année sur 2")
    gestion_CAA.set(str(CAA)+" (2 an/parcelle)")
    gestion_zoxamide.set(str(zoxamide)+" (3 an/parcelle)")
    gestion_Qil.set(str(Qil)+" (1 an/parcelle)")
    gestion_qosi.set(str(qosi)+" (1 an/parcelle)")
    gestion_fluopicolide.set(str(fluopicolide)+" (1 an/parcelle)")
    gestion_oxathiapiproline.set(str(oxathiapiproline)+" (1 an/parcelle)")
    gestion_anilides.set(str(anilides)+" (2 an/parcelle)")
    gestion_cymoxanil.set(str(cymoxanil)+" (2 an/parcelle)")
    gestion_qoicontact.set(str(qoicontact)+" (1 an/parcelle)")
    gestion_spiroxamine.set(str(spiroxamine)+" (2 an/parcelle)")
    gestion_APK.set(str(APK)+" (2 an/parcelle)")
    gestion_SDHI_fluopyram.set(str(SDHI_fluopyram)+" (2 substances≠/an/parcelle)")
    gestion_SDHI_boscalid.set(str(SDHI_boscalid)+" (2 substances≠/an/parcelle)")
    gestion_SDHI_fluxapyroxad.set(str(SDHI_fluxapyroxad)+" (2 substances≠/an/parcelle)")
    gestion_IBS1.set(str(IBS1)+" (1 an/parcelle)")
    gestion_AZN.set(str(AZN)+" (1 an/parcelle)")
    gestion_cyflufenamid.set(str(cyflufenamid)+" (1 an/parcelle)")
    gestion_phenylpyrroles.set(phenylpyrroles)
    gestion_ANP.set(ANP)
    gestion_IBS3.set(IBS3)
    gestion_SDHI.set(SDHI)
    gestion_CAA.set(CAA)
    gestion_zoxamide.set(zoxamide)
    gestion_Qil.set(Qil)
    gestion_qosi.set(qosi)
    gestion_fluopicolide.set(fluopicolide)
    gestion_oxathiapiproline.set(oxathiapiproline)
    gestion_anilides.set(anilides)
    gestion_cymoxanil.set(cymoxanil)
    gestion_qoicontact.set(qoicontact)
    gestion_spiroxamine.set(spiroxamine)
    gestion_APK.set(APK)
    gestion_SDHI_fluopyram.set(SDHI_fluopyram)
    gestion_SDHI_boscalid.set(SDHI_boscalid)
    gestion_SDHI_fluxapyroxad.set(SDHI_fluxapyroxad)
    gestion_IBS1.set(IBS1)
    gestion_AZN.set(AZN)
    gestion_cyflufenamid.set(cyflufenamid)

    return(IFT_classique , IFT_biocontrole , IFT_total , IFT_mildiou , IFT_oidium , IFT_botrytis , IFT_autres_fongicides , IFT_fongicide_total , IFT_confusion_sexuelle , IFT_acaricides , IFT_autres_acaricides , IFT_insecticide_total , IFT_hors_herbicides , IFT_herbicide_prelevee, IFT_herbicide_postlevee, IFT_herbicide, cyflufenamid, AZN, IBS1, SDHI_fluxapyroxad, SDHI_boscalid, SDHI_fluopyram, APK, spiroxamine, qoicontact, cymoxanil, anilides, oxathiapiproline, fluopicolide, qosi, Qil, zoxamide, CAA, SDHI, IBS3, ANP, phenylpyrroles, premiere_date_traitement)

def ouvrir_pdf():
    # Ouvrir = 0 ; Enregistrer = 1
    ouvrir_ou_enregistrer = 0
    # Récupérer les cases cochées :
    cases_cochees = [resume_IFT.get(), resume_gestion_resistances.get() , resume_traitements.get()]

    date_debut = indiquer_date_debut.get()
    date_fin = indiquer_date_fin.get()
    date_debut_formatee = time.strptime(date_debut, "%d/%m/%Y")
    date_fin_formatee = time.strptime(date_fin, "%d/%m/%Y")

    if date_debut_formatee >= date_fin_formatee:
        periode_indiquee = 0
    else :
        periode_indiquee = 1
    nom_parcelle=liste_parcelles_recap.get()
    surface_totale_exploitation = 0
    for parcelle in list(dictionnaire_parcelles.keys()) :
        surface_totale_exploitation=float(surface_totale_exploitation)+float(dictionnaire_parcelles[parcelle])

    nom_fichier=recherche_produit.enregistrer_en_pdf(ouvrir_ou_enregistrer, nom_parcelle,surface_totale_exploitation,periode_indiquee,date_debut,date_fin,cases_cochees) # TODO cases_cochees
    nom_fichier=str(nom_fichier)
    #chemin = "./recapitulatifs/"+str(nom_fichier)
    chemin = "./recapitulatifs/"
    chemin_actuel = os.getcwd()
    os.chdir(chemin)
    #os.system(nom_fichier)
    webbrowser.open_new(nom_fichier)
    os.chdir(chemin_actuel)
    #webbrowser.open_new(nom_fichier)

def enregistrer_pdf():
    # Ouvrir = 0 ; Enregistrer = 1
    ouvrir_ou_enregistrer = 1
    # Récupérer les cases cochées :
    cases_cochees = [resume_IFT.get(), resume_gestion_resistances.get() , resume_traitements.get()]

    date_debut = indiquer_date_debut.get()
    date_fin = indiquer_date_fin.get()
    date_debut_formatee = time.strptime(date_debut, "%d/%m/%Y")
    date_fin_formatee = time.strptime(date_fin, "%d/%m/%Y")

    if date_debut_formatee >= date_fin_formatee:
        periode_indiquee = 0
    else :
        periode_indiquee = 1
    nom_parcelle=liste_parcelles_recap.get()
    surface_totale_exploitation = 0
    for parcelle in list(dictionnaire_parcelles.keys()) :
        surface_totale_exploitation=float(surface_totale_exploitation)+float(dictionnaire_parcelles[parcelle])

    chemin_fichier=recherche_produit.enregistrer_en_pdf(ouvrir_ou_enregistrer, nom_parcelle,surface_totale_exploitation,periode_indiquee,date_debut,date_fin,cases_cochees) # TODO cases_cochees
    chemin_fichier=str(chemin_fichier)
    #chemin = "./recapitulatifs/"+str(nom_fichier)
    #os.system(nom_fichier)
    webbrowser.open_new(chemin_fichier)
    #webbrowser.open_new(nom_fichier)

def update_label_et_superficie():
    parcelle_selectionnee.config(text=str(dictionnaire_parcelles[parcelle_choisie.get()])+" hectare(s)")
    parcelle_selectionnee_scale.set(dictionnaire_parcelles[parcelle_choisie.get()])
    s1.configure(to=float(dictionnaire_parcelles[parcelle_choisie.get()]))
    s1.set(float(dictionnaire_parcelles[parcelle_choisie.get()]))
    update_superficie_a_choisir()
    #calcul_IFT()

def lire_traitements(): # pour onglet Traitements
    tree.delete(*tree.get_children())
    # Add new data in Treeview widget
    traitements = pd.read_excel('IFT.xlsx')
    #traitements = traitements[['date_traitement','nom_produit', '']]
    traitements = traitements.drop(['date','numero_AMM','substances_actives','si_herbicide','passage'], axis=1)
    tree["column"] = list(traitements.columns)
    tree["show"] = "headings"

    # For Headings iterate over the columns
    for col in tree["column"]:
       tree.heading(col, text=col)
       tree.column(col, anchor=CENTER, width=90, stretch=NO)

    # Put Data in Rows
    df_rows = traitements.to_numpy().tolist()
    for row in df_rows:
        tree.insert("", "end", values=row)

def supprimer_le_traitement(): # Pour onglet Traitements
    if tree.focus() == "" :
        messagebox.showinfo("Information","Sélectionnez un traitement à supprimer")
        return
    avertissement=tkinter.messagebox.askyesno(title="Supprimer le traitement", message="Êtes-vous sûr de vouloir supprimer ce traitement de la base de données ?")
    if avertissement == True :
        ligne_en_cours = tree.focus()
        # obtenir l'index de la ligne en cours (pour la supprimer de 'IFT.xlsx'
        index_de_la_ligne = tree.index(ligne_en_cours)
        IFT_openpyxl = openpyxl.reader.excel.load_workbook('IFT.xlsx')
        sheet_en_cours = IFT_openpyxl['IFT']
        sheet_en_cours.delete_rows(index_de_la_ligne+2, 1)
        IFT_openpyxl.save('IFT.xlsx')
        tree.delete(tree.selection()[0])
        infos_IFT_toute_la_base_de_donnees()
        recapitulatif_total()

def changer_surface_traitee() :
    surface_traitee.delete(0,END)
    surface_traitee.insert(0,parcelle_selectionnee_scale.get())
    calcul_IFT()

def changer_quantite_traitee() :
    quantite_traitee.delete(0,END)
    quantite_traitee.insert(0,quantite_selectionnee_scale.get())
    calcul_IFT()


# Création des différents modules (entrée de texte et listes) [je les mets avant d'organiser les fenêtres (grid), car sinon les fonctions associées aux boutons/événements appellent des noms de variables non chargées)
# Onglet "IFT"
# Zone d'entrée de texte par l'utilisateur / liste de sélection par l'utilisateur
nom_produit = Entry(onglet1, width=45, justify="center")                                                                      # Nom du produit entré par l'utilisateur
combobox_liste_produits = ttk.Combobox(onglet1, width = 42, textvariable = produit_choisi)                  # Liste qui contiendra les produits trouvés
frame_utilisation_herbicide = ttk.Frame(onglet1)
combobox_utilisations = ttk.Combobox(frame_utilisation_herbicide, width = 30, textvariable = utilisation_choisie)                      # Liste qui contiendra les doses possibles pour le produit sélectionné
label_dose = tkinter.Label (onglet1 ,textvariable = dose_choisie)
surface_totale = Entry(onglet1, width=42)                                                                   # Surface entrée par l'utilisateur
frame_surface_traitee = ttk.Frame(onglet1)
surface_traitee = Entry(frame_surface_traitee, width=30, justify="center")                                                                  # idem
frame_quantite_traitee = ttk.Frame(onglet1)
quantite_traitee = Entry(frame_quantite_traitee, width=30, justify="center")                                                                 # Quantité de produit entrée par l'utilisateur
frame_parcelles_traitees = ttk.Frame(onglet1)
combobox_parcelles = ttk.Combobox(frame_parcelles_traitees, values=list(dictionnaire_parcelles.keys()), width = 25, textvariable = parcelle_choisie, justify="center")
combobox_herbicide = ttk.Combobox(frame_utilisation_herbicide, values=list(['postlevée','prélevée']), width = 10,textvariable=produit_herbicide)
# Boutons
#bouton_chercher_produit_entre = tkinter.Button (text = "Rechercher" , command = rechercher_produit)         # Bouton : Rechercher le produit entré dans la case "nom_produit"
#bouton_valider_produit_trouve = tkinter.Button (text = "Valider" , command = rechercher_meme_dose)          # Bouton : Valider la recherche pour le produit choisi dans la liste "combobox_liste_produits"
#bouton_valider_produit_cherche = tkinter.Button (text = "Valider" , command = rechercher_doses_unite)       # Bouton : Valider la recherche pour la dose choisie dans la liste "combobox_utilisations"
#bouton_calcul_IFT = tkinter.Button (text = "Valider" , command = calcul_IFT)                                # Bouton : Calculer l'IFT
# Zone de texte (labels)
parcelle_totale = tkinter.Label(onglet1, text="hectare(s)")
zone_texte_nom_produit_a_chercher = tkinter.Label (onglet1 ,text = "Nom du produit cherché :")
zone_texte_entrer_produit_cherche = tkinter.Label (onglet1 ,text = "Sélectionnez le produit :")
zone_texte_selectionner_dose = tkinter.Label (onglet1 ,text = "Sélectionnez l'utilisation :")
frame_IFT_traitement = ttk.Frame(onglet1, borderwidth=5, relief="solid")
zone_texte_IFT = tkinter.Label (frame_IFT_traitement ,text = "IFT du traitement :")
zone_texte_IFT_calcule = tkinter.Label (frame_IFT_traitement ,textvariable = IFT)

#frame_pour_dose_unite_passage = tkinter.Frame(onglet1)
frame_dose_passage = ttk.Frame(onglet1)
frame_infos_dose_unite = ttk.Frame(frame_dose_passage, borderwidth=5, relief="solid")
frame_infos_passage = ttk.Frame(frame_dose_passage, borderwidth=5, relief="solid")
frame_infos_dose_unite_trouvees = ttk.Frame(frame_infos_dose_unite)
frame_label_texte_dose_reglementaire = tkinter.Frame(frame_infos_dose_unite)
zone_texte_dose_reglementaire = tkinter.Label (frame_label_texte_dose_reglementaire ,text = "Dose réglementaire")
zone_texte_unite = tkinter.Label (frame_infos_dose_unite ,text = "Unité")
zone_texte_nombre_passages = tkinter.Label (frame_infos_passage ,text = "Nombre de traitements max.")
zone_texte_dose_reglementaire_trouvee = tkinter.Label (frame_infos_dose_unite_trouvees ,textvariable = dose_autorisee)
zone_texte_unite_trouvee = tkinter.Label (frame_infos_dose_unite_trouvees ,textvariable = unite)
zone_texte_nombre_passages_trouve = tkinter.Label (frame_infos_passage ,textvariable = application, width=10)
zone_texte_surface_totale = tkinter.Label (onglet1 ,text = "Surface totale de la parcelle :")
frame_infos_dose_unite_trouvees2 = ttk.Frame(frame_quantite_traitee)
frame_label_hectare = ttk.Frame(frame_surface_traitee)
zone_texte_hectare = tkinter.Label (frame_label_hectare ,text = "hectare(s)")
zone_texte_unite_trouvee2 = tkinter.Label (frame_infos_dose_unite_trouvees2 ,textvariable = unite, width=7)
zone_texte_surface_traitee = tkinter.Label (onglet1 ,text = "Surface traitée :")
zone_texte_quantite_produit = tkinter.Label (onglet1 ,text = "Dose appliquée :")
zone_texte_passages_restants = tkinter.Label (onglet1, textvariable = label_passages_restants)
# Onglet "Parcelles"



# Organisation des fenêtres et fonctions associées
# Onglet "IFT"
# JUMP LINE
label_vide = tkinter.Label(onglet1, text="")
label_vide.grid(row=0,column=1)
####
    # Ajouter une parcelle

label_ajouter_un_traitement = tkinter.Label(onglet1, text="Ajouter un traitement", font="Times, 18")
label_ajouter_un_traitement.grid(row=1,column=1)


# Zone de recherche du produit
zone_texte_nom_produit_a_chercher.grid(row = 2, column=0)
nom_produit.bind('<KeyRelease>', (lambda event:rechercher_produit()))
nom_produit.grid(row=2,column=1)
nom_produit.focus_force()     # placement du curseur dans la première entrée.
    # Bouton "Rechercher"
#bouton_chercher_produit_entre.grid(row=0,column=2)

# JUMP LINE
label_vide = tkinter.Label(onglet1, text="")
label_vide.grid(row=3,column=1)
####

# Liste des produits trouvés
zone_texte_entrer_produit_cherche.grid(row = 4, column=0)
combobox_liste_produits.bind('<<ComboboxSelected>>', lambda event: rechercher_meme_dose() & rechercher_doses_unite())
combobox_liste_produits['values'] = liste
combobox_liste_produits.grid(row=4,column=1)
    # Bouton "Valider"
#bouton_valider_produit_trouve.grid(row=1,column=2)

# Sélectionner l'utilisation souhaitée (plusieurs possibles, mais toujours "autorisées")
zone_texte_selectionner_dose.grid(row = 5, column=0)
combobox_utilisations.bind('<<ComboboxSelected>>', lambda event: rechercher_doses_unite())
combobox_utilisations.grid(row=0,column=0)
frame_utilisation_herbicide.grid(row=5,column=1)

# JUMP LINE
label_vide = tkinter.Label(onglet1, text="")
label_vide.grid(row=6,column=1)
####

#bouton_valider_produit_cherche.grid(row=2,column=2)

# Afficher dose autorisée , unité , applications_max

zone_texte_dose_reglementaire.grid(row = 0, column=0)
#zone_texte_unite.grid(row = 6, column=1)
zone_texte_dose_reglementaire_trouvee.grid(row=0,column=0)
zone_texte_unite_trouvee.grid(row=0,column=1)
zone_texte_nombre_passages.grid(row = 0, column=0)
zone_texte_nombre_passages_trouve.grid(row=1,column=0)
frame_infos_dose_unite.grid(row=0,column=0)
frame_infos_dose_unite_trouvees.grid(row=1,column=0)
# JUMP COLUMN
label_vide = tkinter.Label(frame_dose_passage, text="")
label_vide.grid(row=0,column=1)
####
frame_infos_passage.grid(row=0,column=2)
frame_label_texte_dose_reglementaire.grid(row=0,column=0)
frame_dose_passage.grid(row=7,column=1)




# JUMP LINE
label_vide = tkinter.Label(onglet1, text="")
label_vide.grid(row=9,column=1)
####

# Demander la parcelle choisie :
zone_texte_surface_totale = tkinter.Label (onglet1 ,text = "Parcelle traitée :")
zone_texte_surface_totale.grid(row = 10, column=0)
# Afficher taille parcelle :
combobox_parcelles.bind('<<ComboboxSelected>>', lambda event: update_label_et_superficie())



combobox_parcelles.grid(row=0,column=0)
frame_parcelles_traitees.grid(row=10,column=1)
parcelle_selectionnee = tkinter.Label(frame_parcelles_traitees, text="")
parcelle_selectionnee.grid(row=0, column=1, sticky="w")
# Si le dictionnaire n'est pas vide, afficher le premier élément, sinon ne rien faire
if dictionnaire_parcelles:
    combobox_parcelles.current(0)
    parcelle_selectionnee.config(text=str(dictionnaire_parcelles[parcelle_choisie.get()])+" hectare(s)")
    surface_traitee.delete(0,END)
    surface_traitee.insert(0,dictionnaire_parcelles[parcelle_choisie.get()])
else:
    combobox_parcelles['values'] = ["Aucune parcelle enregistrée. Allez dans l'onglet 'Parcelles'"] # TODO
    combobox_parcelles.current(0)
    surface_traitee.delete(0,END)


# Demander surface traitée :
zone_texte_surface_traitee.grid(row = 11, column=0)
# SCALE 2022
parcelle_selectionnee_scale=DoubleVar()
if dictionnaire_parcelles:
    parcelle_selectionnee_scale.set(dictionnaire_parcelles[parcelle_choisie.get()])
else :
    parcelle_selectionnee_scale.set(1)

s1 = Scale(frame_surface_traitee, variable=parcelle_selectionnee_scale, resolution=0, from_ = 0.00, to = parcelle_selectionnee_scale.get(), orient = HORIZONTAL, command = lambda event : changer_surface_traitee(), showvalue=0)
s1.grid(row=0,column=0)


def update_les_scales():
    s1.set(surface_traitee.get())
    s2.set(quantite_traitee.get())

surface_traitee.bind('<KeyRelease-comma>', lambda event : changer_virgule_en_point_surface_traitee())
surface_traitee.bind('<KeyRelease>', lambda event: calcul_IFT())
surface_traitee.bind('<FocusOut>', lambda event: update_les_scales())
surface_traitee.grid(row=0,column=1)

zone_texte_unite_trouvee2.grid(row=0,column=1)
frame_surface_traitee.grid(row=11,column=1)
zone_texte_hectare.grid(row=0,column=2)
frame_label_hectare.grid(row=0,column=2)



# Demander quantité produit :
quantite_selectionnee_scale=DoubleVar()
s2 = Scale(frame_quantite_traitee, variable = quantite_selectionnee_scale, resolution=0, from_ = 0.00, to = 1, orient = HORIZONTAL, command = lambda event : changer_quantite_traitee(), showvalue=0)
s2.grid(row=0,column=0)
zone_texte_quantite_produit.grid(row = 12, column=0)
quantite_traitee.bind('<KeyRelease-comma>', lambda event : changer_virgule_en_point_quantite_traitee())
quantite_traitee.bind('<KeyRelease>', lambda event: calcul_IFT())
quantite_traitee.bind('<FocusOut>', lambda event: update_les_scales())
quantite_traitee.grid(row=0,column=1)
frame_infos_dose_unite_trouvees2.grid(row=0,column=2)
frame_quantite_traitee.grid(row=12,column=1)


#bouton_calcul_IFT.grid(row=8,column=1)

# Afficher nombre de passages restant autorisés :
zone_texte_passages_restants.grid(row=8,column=1)

# Demander d'indiquer une date pour le traitement (si rien n'est mis, prendre date du jour)
frame_date_traitement = ttk.Frame(onglet1)
zone_texte_indiquer_une_date=tkinter.Label(frame_date_traitement, text="Date du traitement : ")
zone_texte_indiquer_une_date.grid(row=0,column=0)
indiquer_date = DateEntry(frame_date_traitement, values="Text", year=annee_en_cours, date_pattern="dd/mm/yyyy", locale="fr_FR")
#indiquer_date = DateEntry(onglet1, values="Text", year=2021, state="readonly", date_pattern="dd/mm/yyyy",textvariable=date_traitement)
indiquer_date.grid(row=0, column=1, padx=20, pady=5, sticky=W)
#cal = Calendar(onglet1, selectmode = 'day', year = 2020, month = 5, day = 22)
#cal.grid(row=17,column=1)
frame_date_traitement.grid(row=17,column=1)

# JUMP LINE
label_vide = tkinter.Label(onglet1, text="")
label_vide.grid(row=18,column=1)
####

# AJOUTER ENREGISTREMENT avec le boutons sauvegarde de papa
# Création du bouton de Sauvegarde
sauvegarde = tkinter.Button (onglet1, text = "Enregistrer le traitement" , command = sauvegarde)
sauvegarde.grid(row=19,column=1)
#quitter = tkinter.Button(onglet1, text="Quitter", command=tkquit)
#quitter.grid(row=13,column=2)

# Période date fin (OBLIGE de la mettre ici sinon marche pas... faire avec des classes)
frame_dates_recap = ttk.Frame(onglet4)
indiquer_date_fin = DateEntry(frame_dates_recap, values="Text", year=annee_en_cours, date_pattern="dd/mm/yyyy", locale="fr_FR")
indiquer_date_debut = DateEntry(frame_dates_recap, values="Text", year=annee_en_cours-1, date_pattern="dd/mm/yyyy", locale="fr_FR")
indiquer_date_debut.grid(row=1, column=1, padx=20, pady=5, sticky=W)
label_fleche = tkinter.Label(frame_dates_recap, text=">")
label_fleche.grid(row=1,column=2)
indiquer_date_fin.grid(row=1, column=3, padx=20, pady=5, sticky=W)
label_debut = tkinter.Label(frame_dates_recap, text="Début")
label_debut.grid(row=2,column=1)
label_fin = tkinter.Label(frame_dates_recap, text="Fin")
label_fin.grid(row=2,column=3)

indiquer_date_debut.bind('<<DateEntrySelected>>', lambda event: recapitulatif_total()) # TODO
indiquer_date_fin.bind('<<DateEntrySelected>>', lambda event: recapitulatif_total())
frame_dates_recap.grid(row=3,column=1)

# JUMP LINE
label_vide = tkinter.Label(onglet1, text="")
label_vide.grid(row=20,column=1)
####


# Afficher IFt :
zone_texte_IFT.grid(row=0,column=0)
zone_texte_IFT_calcule.grid(row=0,column=1)
frame_IFT_traitement.grid(row=21,column=1)


# Onglet "Parcelles"

# JUMP LINE
label_vide = tkinter.Label(onglet2, text="")
label_vide.grid(row=0,column=1)
####

    # Ajouter une parcelle
label_ajouter_une_parcelle = tkinter.Label(onglet2, text="Ajouter une parcelle", font="Times, 18")
label_ajouter_une_parcelle.grid(row=1,column=1)

# JUMP LINE
label_vide = tkinter.Label(onglet2, text="")
label_vide.grid(row=2,column=1)
####

label_entrer_un_nom_de_parcelle = tkinter.Label(onglet2, text="Nom de la parcelle : ")
label_entrer_un_nom_de_parcelle.grid(row=3,column=0)
entrer_un_nom_de_parcelle = Entry(onglet2, width=50)
entrer_un_nom_de_parcelle.grid(row=3,column=1)
label_entrer_taille_de_parcelle = tkinter.Label(onglet2, text="Superficie en hectares (ex : 0.2) : ")
label_entrer_taille_de_parcelle.grid(row=4,column=0)
entrer_taille_de_parcelle = Entry(onglet2, width=50)
entrer_taille_de_parcelle.bind('<KeyRelease-comma>', lambda event : changer_virgule_en_point_entrer_taille_de_parcelle())
entrer_taille_de_parcelle.grid(row=4,column=1)

label_superficie_confusion_sexuelle = tkinter.Label(onglet2, text="Superficie en confusion sexuelle : ")
label_superficie_confusion_sexuelle.grid(row=5,column=0)
entrer_superficie_confusion = Entry(onglet2, width=50,textvariable=superficie_confusion_sexuelle)
entrer_superficie_confusion.bind('<KeyRelease-comma>', lambda event : changer_virgule_en_point_entrer_superficie_confusion())
entrer_superficie_confusion.grid(row=5,column=1)

ajouter = tkinter.Button(onglet2, text = "Ajouter cette parcelle" , command = lambda : ajouter_parcelle())
ajouter.grid(row=6,column=1)


# JUMP LINE
label_vide = tkinter.Label(onglet2, text="")
label_vide.grid(row=7,column=1)
####

    # Suprimer une parcelle
label_supprimer_une_parcelle = tkinter.Label(onglet2, text="Supprimer une parcelle", font="Times, 18")
label_supprimer_une_parcelle.grid(row=8,column=1)

# JUMP LINE
label_vide = tkinter.Label(onglet2, text="")
label_vide.grid(row=9,column=1)
####

label_parcelles_de_la_base = tkinter.Label(onglet2, text="Parcelle à supprimer :")
label_parcelles_de_la_base.grid(row=10,column=0)
liste_parcelles_de_la_base = ttk.Combobox(onglet2, values=list(dictionnaire_parcelles.keys()), width = 45, textvariable = parcelle_a_supprimer)
liste_parcelles_de_la_base.bind('<<ComboboxSelected>>', lambda event: infos_IFT_par_parcelle())
if dictionnaire_parcelles:
    liste_parcelles_de_la_base.current(0)
    infos_IFT_par_parcelle()
else:
    liste_parcelles_de_la_base['values'] = [""]
    liste_parcelles_de_la_base.current(0)
liste_parcelles_de_la_base.grid(row=10,column=1)

supprimer = tkinter.Button (onglet2, text = "Supprimer" , command = lambda : supprimer_parcelle())
supprimer.grid(row=11,column=1)


#label_IFT_total_parcelle = tkinter.Label(onglet2, text="IFT de la parcelle : ")
#label_IFT_total_parcelle.grid(row=5,column=1)
#infos_IFT_par_parcelle()    # fonction peut être gourmande ? (car se relance à chaque fois ?)
#
#resultat_IFT_total_parcelle = tkinter.Label(onglet2, textvariable=IFT_total_parcelle)
#resultat_IFT_total_parcelle.grid(row=6,column=1)
#
#label_IFT_total_BDD = tkinter.Label(onglet2, text="IFT sur l'ensemble des parcelles : ")
#label_IFT_total_BDD.grid(row=7,column=1)
#infos_IFT_toute_la_base_de_donnees()    # fonction peut être gourmande ? (car se relance à chaque fois ?)
#
#resultat_IFT_total_BDD = tkinter.Label(onglet2, textvariable=IFT_base_de_donnees)
#resultat_IFT_total_BDD.grid(row=8,column=1)


# Onglet "Update"

ajouter = tkinter.Button(onglet5, text = "Mettre à jour la BDD" , command = lambda : mettre_a_jour_bdd())
#ajouter.grid(row=2,column=1)
label_entrer_un_nom_de_parcelle = tkinter.Label(onglet5, textvariable=resultat_mise_a_jour)
#label_entrer_un_nom_de_parcelle.grid(row=3,column=1)


# Onglet "Récapitulatif"

# JUMP LINE
label_vide = tkinter.Label(onglet4, text="")
label_vide.grid(row=0,column=0)
####

label_ajouter_un_traitement = tkinter.Label(onglet4, text="Obtenir un récapitulatif", font="Times, 18")
label_ajouter_un_traitement.grid(row=1,column=1)

label_recap = tkinter.Label(onglet4, text="Pour quelle parcelle voulez-vous un récapitulatif ?")
label_recap.grid(row=2,column=0)
#### Faire une combobox avec "Exploitation totale" et "par parcelle"
#liste_parcelles_pour_recap = ttk.Combobox(onglet4, values=list(dictionnaire_parcelles.keys())+['-Exploitation'], width = 45, textvariable=liste_parcelles_recap)
liste_parcelles_pour_recap = ttk.Combobox(onglet4, values=["Toute l'exploitation"]+list(dictionnaire_parcelles.keys()), width = 25, textvariable=liste_parcelles_recap)
#if dictionnaire_parcelles:
#    recapitulatif_total() # TODO
liste_parcelles_pour_recap.current(0)
liste_parcelles_pour_recap.bind('<<ComboboxSelected>>', lambda event: recapitulatif_total())
try :
    recapitulatif_total()
except :
    print("Impossible de charger le récapitulatif rapide.")
liste_parcelles_pour_recap.grid(row=2,column=1)

indiquer_une_periode=tkinter.Label(onglet4, text="Sélectionnez la période souhaitée (début > fin) : ") # TODO
indiquer_une_periode.grid(row=3,column=0)





frame_enregistrement = ttk.Frame(onglet4)
ouvrir_recap_en_pdf = tkinter.Button(frame_enregistrement, text = "Ouvrir le récapitulatif" , command = lambda : ouvrir_pdf())
ouvrir_recap_en_pdf.grid(row=0,column=0)

enregistrer_recap_en_pdf = tkinter.Button(frame_enregistrement, text = "Enregistrer le récapitulatif" , command = lambda : enregistrer_pdf())
enregistrer_recap_en_pdf.grid(row=1,column=0)

#indiquer_date_debut = DateEntry(onglet4, values="Text", year=2021, date_pattern="dd/mm/yyyy", locale="fr_FR")
#indiquer_date_debut.grid(row=1, column=1, padx=20, pady=5, sticky=W)
#indiquer_date_fin = DateEntry(onglet4, values="Text", year=2021, date_pattern="dd/mm/yyyy", locale="fr_FR")
#indiquer_date_fin.grid(row=1, column=2, padx=20, pady=5, sticky=W)

# JUMP LINE
label_vide = tkinter.Label(frame_enregistrement, text="")
label_vide.grid(row=2,column=0)
####

# Cases à cocher pour enregistrement :
frame_checkbuttons = ttk.Frame(frame_enregistrement, borderwidth=5, relief="solid")
label_infos_a_inclure = tkinter.Label(frame_checkbuttons, text="Informations à inclure dans le récapitulatif :")
label_infos_a_inclure.grid(row=0,column=0)
Checkbutton(frame_checkbuttons , text="résumé IFT", variable=resume_IFT).grid(row=1, column=0,sticky="w")
Checkbutton(frame_checkbuttons , text="gestion des résistances", variable=resume_gestion_resistances).grid(row=2, column=0, sticky="w")
Checkbutton(frame_checkbuttons , text="résumé des traitements", variable=resume_traitements).grid(row=3, column=0, sticky="w")
frame_checkbuttons.grid(row=3,column=0)
frame_enregistrement.grid(row=5,column=1)



# IFT total
frame_IFT = ttk.Frame(onglet4, borderwidth=5, relief="solid")
frame_header_IFT = ttk.Frame(frame_IFT)
label_header = tkinter.Label(frame_header_IFT, text="Récapitulatif rapide IFT", font='Arial 11')
label_header.grid(row=0,column=0)
frame_header_IFT.grid(row=1,column=0)
label_IFT_total = tkinter.Label(frame_IFT, text="IFT total")
label_IFT_total.grid(row=2,column=0,sticky="w")
resultat_IFT_total = tkinter.Label(frame_IFT, textvariable=IFT_total_recap)
resultat_IFT_total.grid(row=2,column=1)

label_IFT_classique = tkinter.Label(frame_IFT, text="IFT classique")
label_IFT_classique.grid(row=3,column=0,sticky="w")
resultat_IFT_classique = tkinter.Label(frame_IFT, textvariable=IFT_classique_recap)
resultat_IFT_classique.grid(row=3,column=1)

label_IFT_biocontrole = tkinter.Label(frame_IFT, text="IFT biocontrôle")
label_IFT_biocontrole.grid(row=4,column=0,sticky="w")
resultat_IFT_biocontrole = tkinter.Label(frame_IFT, textvariable=IFT_biocontrole_recap)
resultat_IFT_biocontrole.grid(row=4,column=1)

label_IFT_mildiou = tkinter.Label(frame_IFT, text="IFT Mildiou")
label_IFT_mildiou.grid(row=5,column=0,sticky="w")
resultat_IFT_mildiou = tkinter.Label(frame_IFT, textvariable=IFT_mildiou_recap)
resultat_IFT_mildiou.grid(row=5,column=1)

label_IFT_oidium = tkinter.Label(frame_IFT, text="IFT Oïdium")
label_IFT_oidium.grid(row=6,column=0,sticky="w")
resultat_IFT_oidium = tkinter.Label(frame_IFT, textvariable=IFT_oidium_recap)
resultat_IFT_oidium.grid(row=6,column=1)

label_IFT_botrytis = tkinter.Label(frame_IFT, text="IFT Botrytis")
label_IFT_botrytis.grid(row=7,column=0,sticky="w")
resultat_IFT_botrytis = tkinter.Label(frame_IFT, textvariable=IFT_botrytis_recap)
resultat_IFT_botrytis.grid(row=7,column=1)

label_IFT_autres_fongicides = tkinter.Label(frame_IFT, text="IFT autres fongicides")
label_IFT_autres_fongicides.grid(row=8,column=0,sticky="w")
resultat_IFT_autres_fongicides = tkinter.Label(frame_IFT, textvariable=IFT_autres_fongicides_recap)
resultat_IFT_autres_fongicides.grid(row=8,column=1)

label_IFT_fongicide_total = tkinter.Label(frame_IFT, text="IFT fongicides total")
label_IFT_fongicide_total.grid(row=9,column=0,sticky="w")
resultat_IFT_fongicide_total = tkinter.Label(frame_IFT, textvariable=IFT_fongicides_total_recap)
resultat_IFT_fongicide_total.grid(row=9,column=1)

label_IFT_confu = tkinter.Label(frame_IFT, text="IFT confusion sexuelle")
label_IFT_confu.grid(row=10,column=0,sticky="w")
resultat_IFT_confu = tkinter.Label(frame_IFT, textvariable=IFT_confusion_sexuelle_recap)
resultat_IFT_confu.grid(row=10,column=1)

label_IFT_acaricides = tkinter.Label(frame_IFT, text="IFT acariens")
label_IFT_acaricides.grid(row=11,column=0,sticky="w")
resultat_IFT_acaricides = tkinter.Label(frame_IFT, textvariable=IFT_acariens_recap)
resultat_IFT_acaricides.grid(row=11,column=1)

label_IFT_autres_acaricides = tkinter.Label(frame_IFT, text="IFT autres insecticides")
label_IFT_autres_acaricides.grid(row=12,column=0,sticky="w")
resultat_IFT_autres_acaricides = tkinter.Label(frame_IFT, textvariable=IFT_autres_acaricides_recap)
resultat_IFT_autres_acaricides.grid(row=12,column=1)

label_IFT_insecticide_total = tkinter.Label(frame_IFT, text="IFT insecticide total")
label_IFT_insecticide_total.grid(row=13,column=0,sticky="w")
resultat_IFT_insecticide_total = tkinter.Label(frame_IFT, textvariable=IFT_insecticide_recap)
resultat_IFT_insecticide_total.grid(row=13,column=1)

label_IFT_hors_herbicide = tkinter.Label(frame_IFT, text="IFT hors herbicides")
label_IFT_hors_herbicide.grid(row=14,column=0,sticky="w")
resultat_IFT_hors_herbicide = tkinter.Label(frame_IFT, textvariable=IFT_hors_herbicides_recap)
resultat_IFT_hors_herbicide.grid(row=14,column=1)

label_IFT_herbicide_prelevee = tkinter.Label(frame_IFT, text="IFT herbicides prélevée")
label_IFT_herbicide_prelevee.grid(row=15,column=0,sticky="w")
resultat_IFT_herbicide_prelevee = tkinter.Label(frame_IFT, textvariable=IFT_herbicides_prelevee_recap)
resultat_IFT_herbicide_prelevee.grid(row=15,column=1)

label_IFT_herbicide_postlevee = tkinter.Label(frame_IFT, text="IFT herbicides postlevée")
label_IFT_herbicide_postlevee.grid(row=16,column=0,sticky="w")
resultat_IFT_herbicide_postlevee = tkinter.Label(frame_IFT, textvariable=IFT_herbicides_postlevee_recap)
resultat_IFT_herbicide_postlevee.grid(row=16,column=1)

label_IFT_herbicide = tkinter.Label(frame_IFT, text="IFT herbicides")
label_IFT_herbicide.grid(row=17,column=0,sticky="w")
resultat_IFT_herbicide = tkinter.Label(frame_IFT, textvariable=IFT_herbicides_recap)
resultat_IFT_herbicide.grid(row=17,column=1)

frame_IFT.grid(row=5,column=0)


# Gestion des résistances
#label_gestion_phenylpyrroles = tkinter.Label(frame_IFT, text="Phénylpyrolles")
#label_gestion_phenylpyrroles.grid(row=2,column=2)
resultat_gestion_phenylpyrroles = tkinter.Label(frame_IFT, textvariable=gestion_phenylpyrroles)
#resultat_gestion_phenylpyrroles.grid(row=2,column=3)
#
#label_gestion_ANP = tkinter.Label(frame_IFT, text="ANP")
#label_gestion_ANP.grid(row=3,column=2)
resultat_gestion_ANP = tkinter.Label(frame_IFT, textvariable=gestion_ANP)
#resultat_gestion_ANP.grid(row=3,column=3)
#
#label_gestion_IBS3 = tkinter.Label(frame_IFT, text="IBS3")
#label_gestion_IBS3.grid(row=4,column=2)
resultat_gestion_IBS3 = tkinter.Label(frame_IFT, textvariable=gestion_IBS3)
#resultat_gestion_IBS3.grid(row=4,column=3)
#
#label_gestion_SDHI = tkinter.Label(frame_IFT, text="SDHI")
#label_gestion_SDHI.grid(row=5,column=2)
resultat_gestion_SDHI = tkinter.Label(frame_IFT, textvariable=gestion_SDHI)
#resultat_gestion_SDHI.grid(row=5,column=3)
#
#label_gestion_CAA = tkinter.Label(frame_IFT, text="CAA")
#label_gestion_CAA.grid(row=6,column=2)
resultat_gestion_CAA = tkinter.Label(frame_IFT, textvariable=gestion_CAA)
#resultat_gestion_CAA.grid(row=6,column=3)
#
#label_gestion_zoxamide = tkinter.Label(frame_IFT, text="Zoxamide")
#label_gestion_zoxamide.grid(row=7,column=2)
resultat_gestion_zoxamide = tkinter.Label(frame_IFT, textvariable=gestion_zoxamide)
#resultat_gestion_zoxamide.grid(row=7,column=3)
#
#label_gestion_Qil = tkinter.Label(frame_IFT, text="Qil")
#label_gestion_Qil.grid(row=8,column=2)
resultat_gestion_Qil = tkinter.Label(frame_IFT, textvariable=gestion_Qil)
#resultat_gestion_Qil.grid(row=8,column=3)
#
#label_gestion_qosi = tkinter.Label(frame_IFT, text="QoSI/Qiol")
#label_gestion_qosi.grid(row=9,column=2)
resultat_gestion_qosi = tkinter.Label(frame_IFT, textvariable=gestion_qosi)
#resultat_gestion_qosi.grid(row=9,column=3)
#
#label_gestion_fluopicolide = tkinter.Label(frame_IFT, text="Fluopicolide")
#label_gestion_fluopicolide.grid(row=10,column=2)
resultat_gestion_fluopicolide = tkinter.Label(frame_IFT, textvariable=gestion_fluopicolide)
#resultat_gestion_fluopicolide.grid(row=10,column=3)
#
#label_gestion_oxathiapiproline = tkinter.Label(frame_IFT, text="Oxathiapiproline")
#label_gestion_oxathiapiproline.grid(row=11,column=2)
resultat_gestion_oxathiapiproline = tkinter.Label(frame_IFT, textvariable=gestion_oxathiapiproline)
#resultat_gestion_oxathiapiproline.grid(row=11,column=3)
#
#label_gestion_anilides = tkinter.Label(frame_IFT, text="Anilides")
#label_gestion_anilides.grid(row=12,column=2)
resultat_gestion_anilides = tkinter.Label(frame_IFT, textvariable=gestion_anilides)
#resultat_gestion_anilides.grid(row=12,column=3)
#
#label_gestion_cymoxanil = tkinter.Label(frame_IFT, text="Cymoxanil")
#label_gestion_cymoxanil.grid(row=13,column=2)
resultat_gestion_cymoxanil = tkinter.Label(frame_IFT, textvariable=gestion_cymoxanil)
#resultat_gestion_cymoxanil.grid(row=13,column=3)
#
#label_gestion_qoicontact = tkinter.Label(frame_IFT, text="QoI+contact")
#label_gestion_qoicontact.grid(row=14,column=2)
resultat_gestion_qoicontact = tkinter.Label(frame_IFT, textvariable=gestion_qoicontact)
#resultat_gestion_qoicontact.grid(row=14,column=3)
#
#label_gestion_spiroxamine = tkinter.Label(frame_IFT, text="Spiroxamine")
#label_gestion_spiroxamine.grid(row=15,column=2)
resultat_gestion_spiroxamine = tkinter.Label(frame_IFT, textvariable=gestion_spiroxamine)
#resultat_gestion_spiroxamine.grid(row=15,column=3)
#
#label_gestion_APK = tkinter.Label(frame_IFT, text="APK")
#label_gestion_APK.grid(row=16,column=2)
resultat_gestion_APK = tkinter.Label(frame_IFT, textvariable=gestion_APK)
#resultat_gestion_APK.grid(row=16,column=3)
#
#label_gestion_SDHI_fluopyram = tkinter.Label(frame_IFT, text="SDHI (fluopyram)")
#label_gestion_SDHI_fluopyram.grid(row=17,column=2)
resultat_gestion_SDHI_fluopyram = tkinter.Label(frame_IFT, textvariable=gestion_SDHI_fluopyram)
#resultat_gestion_SDHI_fluopyram.grid(row=17,column=3)
#
#label_gestion_SDHI_boscalid = tkinter.Label(frame_IFT, text="SDHI (boscalid)")
#label_gestion_SDHI_boscalid.grid(row=18,column=2)
resultat_gestion_SDHI_boscalid = tkinter.Label(frame_IFT, textvariable=gestion_SDHI_boscalid)
#resultat_gestion_SDHI_boscalid.grid(row=18,column=3)
#
#label_gestion_SDHI_fluxapyroxad = tkinter.Label(frame_IFT, text="SDHI (fluxapyroxad)")
#label_gestion_SDHI_fluxapyroxad.grid(row=19,column=2)
resultat_gestion_SDHI_fluxapyroxad = tkinter.Label(frame_IFT, textvariable=gestion_SDHI_fluxapyroxad)
#resultat_gestion_SDHI_fluxapyroxad.grid(row=19,column=3)
#
#label_gestion_IBS1 = tkinter.Label(frame_IFT, text="IBS1")
#label_gestion_IBS1.grid(row=20,column=2)
resultat_gestion_IBS1 = tkinter.Label(frame_IFT, textvariable=gestion_IBS1)
#resultat_gestion_IBS1.grid(row=20,column=3)
#
#label_gestion_AZN = tkinter.Label(onglet4, text="AZN")
#label_gestion_AZN.grid(row=21,column=2)
resultat_gestion_AZN = tkinter.Label(frame_IFT, textvariable=gestion_AZN)
#resultat_gestion_AZN.grid(row=21,column=3)
#
#label_gestion_cyflufenamid = tkinter.Label(frame_IFT, text="Cyflufenamid")
#label_gestion_cyflufenamid.grid(row=22,column=2)
resultat_gestion_cyflufenamid = tkinter.Label(frame_IFT, textvariable=gestion_cyflufenamid)
#resultat_gestion_cyflufenamid.grid(row=22,column=3)

# return(IFT_classique , IFT_biocontrole , IFT_total , IFT_mildiou , IFT_oidium , IFT_botrytis , IFT_autres_fongicides , IFT_fongicide_total , IFT_confusion_sexuelle , IFT_acaricides , IFT_autres_acaricides , IFT_insecticide_total , IFT_hors_herbicides , IFT_herbicide)


########################
# Onglet 'Traitements' #
########################

# Create a Treeview widget
#tree = ttk.Treeview(onglet3,selectmode ='extended')
tree = ttk.Treeview(onglet3,selectmode ='browse', height=24)
lire_traitements()

# Barre verticale (mais problemes... TODO)
scrollbar_verticale = ttk.Scrollbar(onglet3, orient="vertical", command=tree.yview)
scrollbar_verticale.pack(side='right',fill='y')
#tree.configure(yscroll=scrollbar_verticale.set)
tree.configure(yscrollcommand=scrollbar_verticale.set)
tree.pack()

#tree.bind("<<TreeviewSelect>>", lambda event : obtenir_le_nom_de_la_ligne())

#tree.grid(row=0,column=0)
#tree.pack(expand=YES, fill="both")
scrollbar_horizontale = ttk.Scrollbar(onglet3, orient="horizontal", command=tree.xview)
tree.configure(xscroll=scrollbar_horizontale.set)
#scrollbar_horizontale.grid(row=1, column=0, sticky='nsew')
scrollbar_horizontale.pack(fill='x')


bouton_supprimer_traitement = tkinter.Button (onglet3, text = "Supprimer le traitement" , command = lambda : supprimer_le_traitement() )
bouton_supprimer_traitement.pack()



# Bar de menu
menubar = Menu(fenetre)
file = Menu(menubar, tearoff=0)
file.add_command(label="Ouvrir un récapitulatif", command = lambda : onglets.select(onglet4) & ouvrir_pdf())
file.add_command(label="Enregistrer un récapitulatif...", command = lambda : onglets.select(onglet4) & enregistrer_pdf())
# -------
file.add_separator()
# -------
file.add_command(label="Quitter", command=fenetre.quit)
menubar.add_cascade(label="Fichier", menu=file)

edit = Menu(menubar, tearoff=0)
edit.add_command(label="Ajouter un traitement", command = lambda : onglets.select(onglet1))
edit.add_command(label="Liste des traitements", command = lambda : onglets.select(onglet3))
# -------
edit.add_separator()
edit.add_command(label="Ajouter/supprimer une parcelle", command = lambda : onglets.select(onglet2))
# -------
edit.add_separator()
edit.add_command(label="Mettre la base des produits à jour", command = lambda : mettre_a_jour_bdd())

menubar.add_cascade(label="Édition", menu=edit)
help = Menu(menubar, tearoff=0)
help.add_command(label="À propos", command = lambda : a_propos())
menubar.add_cascade(label="Aide", menu=help)

fenetre.config(menu=menubar)



fenetre.mainloop ()
